import json
from typing import List
import openpyxl
from openpyxl.drawing.image import Image
import cv2
import pandas as pd
import matplotlib.pyplot as plt
import os
import math
import numpy as np
from PIL import Image
import time
from prompt import prompt_critic_system, prompt_critic_user
import re

def extract_answer(text):
    # Attempt to extract using standard format
    text = text.replace("```json", "```")
    match = re.search(r'```(.*?)```', text, re.DOTALL)
    
    # If paired ``` markers are found, return the content between them
    if match:
        matched_text = match.group(1).strip()
        return matched_text
    
    # If no paired ``` markers are found, return the entire string
    return text.replace("```", "").replace("}}", "}").replace("{{", "{").strip()




def read_json(rpath: str):
    with open(rpath, 'r', encoding='utf-8') as file:
        data = json.load(file)
        return data


def write_json(anns: List, wpath: str):
    json.dump(anns, open(wpath, "w"))




def add_visilize2screenshot(image_rpath, ann, tag):
    """Add visualization markers to screenshots for the mind2web dataset"""

    if type(ann) == dict:
        # Process mind2web format actions
        if "operation" not in ann or ann["operation"]["original_op"] not in ['CLICK', 'TYPE', 'SELECT', 'HOVER', 'ENTER']:
            return image_rpath
        
        # Calculate click center point
        bbox = ann.get("bbox", {})
        if not bbox:
            return image_rpath
            
        click_point = [
            bbox["x"] + bbox["width"]/2,
            bbox["y"] + bbox["height"]/2
        ]
    else:

        try:
            ann = json.loads(ann)
        except Exception as e:
            # print(f"### json.loads error: {e}")
            return image_rpath
        # Process class object format actions
        if ann["action_type"] not in ["click", "type", "select"]:
            return image_rpath

        click_point = ann["click_point"]
        
        # Read image to get dimensions
        if isinstance(image_rpath, str):
            image = cv2.imread(image_rpath)
            height, width = image.shape[:2]
        elif isinstance(image_rpath, Image.Image):
            width, height = image_rpath.size
        else:
            raise ValueError("image_rpath must be a str or PIL.Image.Image")
            

    image = cv2.imread(image_rpath)
    height, width, _ = image.shape

    x = int(float(click_point[0]))
    y = int(float(click_point[1]))

    overlay = image.copy()
    cv2.circle(overlay, (x, y), 20, (0, 0, 255), -1)
    alpha = 0.5  # 50% transparency
    cv2.addWeighted(overlay, alpha, image, 1 - alpha, 0, image)

    image_wpath = image_rpath.split(".jpg")[0] + f"_{tag}.png"
    cv2.imwrite(image_wpath, image) 

    return image_wpath.replace("\\", "/")


    

def write_to_excel(anns, wpath):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="image")
    ws.cell(row=1, column=2, value="image(add point)")
    ws.cell(row=1, column=3, value="task")
    ws.cell(row=1, column=4, value="history action")
    ws.cell(row=1, column=5, value="current action")
    ws.cell(row=1, column=6, value="rating")
    ws.cell(row=1, column=7, value="explanation")

    for idx, ann in enumerate(anns, start=2):
        ws.cell(row=idx, column=3, value=ann["task"])
        ws.cell(row=idx, column=4, value="\n".join(ann["action_desc_list"]))
        ws.cell(row=idx, column=5, value=ann["action_desc_list"][ann["step_id"]])
        ws.cell(row=idx, column=6, value=ann["rating"])
        ws.cell(row=idx, column=7, value=ann["explanation"])

        img = Image(ann["image_list"][ann["step_id"]].replace("\\", "/"))
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'A{idx}')
        img = Image(ann["add_point_image_list"][ann["step_id"]])
        img.width, img.height = (240, 480)
        ws.row_dimensions[idx].height = 400
        ws.add_image(img, f'B{idx}')


    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    wb.save(wpath)


def parse_response(response):
    try:
        response = response.replace("```", "").replace("json", "")
        return json.loads(response)
    except:
        return -1
    

def write_jsonl(anns, wpath):
    with open(wpath, 'w', encoding='utf - 8') as f:
        for item in anns:
            json_line = json.dumps(item)
            f.write(json_line + '\n')


def read_jsonl(rpath):
    data = []
    with open(rpath, 'r', encoding='utf - 8') as f:
        for idx, line in enumerate(f):
            line = line.strip()
            try:
                json_obj = json.loads(line)
                data.append(json_obj)
            except:
                print(f"Error decoding JSON on line: {idx}")
    return data


def read_xlsx(rpath):
    data = pd.read_excel(rpath)
    return data.to_dict(orient="records")


def dict_mean(dict_list):
    mean_dict = {}
    if len(dict_list) > 0:
        for key in dict_list[0].keys():
            if "min" in key:
                mean_dict[key] = min(d[key] for d in dict_list)
            elif "max" in key:
                mean_dict[key] = max(d[key] for d in dict_list)
            else:
                mean_dict[key] = sum(d[key] for d in dict_list) / len(dict_list)
    return mean_dict


def smooth(scalars: List[float]) -> List[float]:
    if len(scalars) == 0:
        return []

    last = scalars[0]
    smoothed = []
    weight = 1.9 * (1 / (1 + math.exp(-0.05 * len(scalars))) - 0.5)  # a sigmoid function
    for next_val in scalars:
        smoothed_val = last * weight + (1 - weight) * next_val
        smoothed.append(smoothed_val)
        last = smoothed_val
    return smoothed


def plot_loss(log_dir: str, keys: List[str] = ["loss"]) -> None:
    plt.switch_backend("agg")
    data = read_jsonl(os.path.join(log_dir, "train_log.jsonl"))

    for key in keys:
        steps, metrics = [], []
        for i in range(len(data)):
            if key in data[i]:
                steps.append(data[i]["step"])
                metrics.append(data[i][key])

        plt.figure()
        plt.plot(steps, metrics, color="#1f77b4", alpha=0.4, label="original")
        plt.plot(steps, smooth(metrics), color="#1f77b4", label="smoothed")
        plt.title(f"{key} of {log_dir}")
        plt.xlabel("step")
        plt.ylabel(key)
        plt.legend()
        figure_path = os.path.join(log_dir, "training_{}.png".format(key))
        plt.savefig(figure_path, format="png", dpi=100)
        print("Figure saved at:", figure_path)


def update_trajectory(anns, results):
    assert len(results) == len(anns)

    for (result, ann) in zip(results, anns):
        history_action_desc = ""
        for i in range(ann["step_id"] - 1):
            history_action_desc += f"step {i}: {ann['action_desc_list'][i]}\n"
        
        ann["critic_input"] = prompt_critic_system + prompt_critic_user.format(ann["task"], history_action_desc, result["output"])
        ann["policy_output"] = result["output"]
        try:
            output = extract_answer(result["output"])
            ann["critic_image"] = add_visilize2screenshot(ann["policy_image"], output, "policy")
        except Exception as e:
            print(f"### update_trajectory error: {e}")
            ann["critic_image"] = ann["policy_image"]
        

    return anns


def resize_image(img, max_size=1024, save_path=None):
    """Scale an image proportionally to ensure the longest side does not exceed max_size, optionally save to a specified path."""
    if isinstance(img, str):
        img = Image.open(img)
    
    width, height = img.size
    if max(width, height) <= max_size:
        resized_img = img
    else:
        # Calculate scaling ratio
        ratio = max_size / max(width, height)
        new_width = int(width * ratio)
        new_height = int(height * ratio)
        
        # Resize the image
        resized_img = img.resize((new_width, new_height), Image.LANCZOS)
    
    # Save the image if a save path is specified
    if save_path:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        resized_img.save(save_path)
        return save_path
    
    return resized_img

def get_first_number_or_default(s):
    match = re.search(r'\d+', s)
    if match:
        num = int(match.group())
        if num == 1 or num == 2:
            return num
    return 1